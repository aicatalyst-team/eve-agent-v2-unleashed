"""LLM code generation for FlowState workflow orchestrator."""

import os
import json
from typing import Dict, Any, List
import ollama
from flowstate.nodes import Node, NodeType


import re as _re


def _py_sanitize(code: str) -> str:
    """Replace JSON literals that LLMs sometimes emit with valid Python."""
    # Only replace whole-word occurrences to avoid touching string contents
    code = _re.sub(r'\bnull\b', 'None', code)
    code = _re.sub(r'\btrue\b', 'True', code)
    code = _re.sub(r'\bfalse\b', 'False', code)
    return code


# ── Built-in node implementations ────────────────────────────────────────────
# Keyed by node title. Each callable receives the Node and returns a code string.
# These run before LLM generation so common nodes always produce working code.

def _impl_file_trigger(node):
    path = node.properties.get("path", "")
    event = node.properties.get("event", "modified")
    return f'''def file__trigger():
    """Return the configured file path for downstream nodes."""
    file_path = {repr(path)}.strip().strip('"').strip("'")
    print(f"File Trigger: watching {{repr(file_path)}} for {repr(event)} events")
    return {{"file_path": file_path}}'''


def _impl_excel_reader(node):
    sheet = node.properties.get("sheet_name", "Sheet1")
    return f'''def excel__reader(file_path):
    """Read all rows from an Excel (.xlsx) or CSV file into a list of dicts."""
    import os as _os
    if not file_path:
        print("Excel Reader: no file_path provided")
        return {{"data": []}}
    file_path = file_path.strip().strip('"').strip("'")
    ext = _os.path.splitext(file_path)[1].lower()

    def _read_as_csv(path):
        import csv
        with open(path, newline="", encoding="utf-8-sig") as _f:
            return [dict(row) for row in csv.DictReader(_f)]

    if ext == ".csv":
        data = _read_as_csv(file_path)
    else:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet_name = {repr(sheet)}
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(v is not None for v in row):
                    data.append(dict(zip(headers, row)))
        except Exception:
            # File may be CSV renamed to .xlsx — fall back to CSV reader
            data = _read_as_csv(file_path)
    print(f"Excel Reader: read {{len(data)}} rows from {{repr(file_path)}}")
    return {{"data": data}}'''


def _impl_file_output(node):
    fmt = node.properties.get("format", "json")
    default_path = node.properties.get("path", "").strip().strip('"').strip("'") or "output.json"
    return f'''def file__output(data, file_path):
    """Write data to a file (json or txt)."""
    import json as _json
    out_path = (file_path or {repr(default_path)}).strip().strip('"').strip("'")
    fmt = {repr(fmt)}
    try:
        if fmt in ("json", ""):
            with open(out_path, "w", encoding="utf-8") as _f:
                _json.dump(data, _f, indent=2, default=str)
        else:
            with open(out_path, "w", encoding="utf-8") as _f:
                if isinstance(data, (list, dict)):
                    _f.write(_json.dumps(data, indent=2, default=str))
                else:
                    _f.write(str(data))
        print(f"File Output: wrote to {{repr(out_path)}}")
    except Exception as _e:
        print(f"File Output error: {{_e}}")
    return {{}}'''


def _impl_email_sender(node):
    smtp = node.properties.get("smtp_server", "smtp.gmail.com")
    port = node.properties.get("port", 587)
    return f'''def email__sender(recipient, subject, body):
    """Send an email via SMTP."""
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    smtp_server = {repr(smtp)}
    port = {port}
    sender = os.getenv("EMAIL_USER", "")
    password = os.getenv("EMAIL_PASS", "")
    msg = MIMEMultipart()
    msg["From"] = sender
    msg["To"] = recipient
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))
    try:
        with smtplib.SMTP(smtp_server, port) as server:
            server.starttls()
            if sender and password:
                server.login(sender, password)
            server.send_message(msg)
        print(f"Email sent to {{repr(recipient)}}")
        return {{"status": True}}
    except Exception as _e:
        print(f"Email error: {{_e}}")
        return {{"status": False}}'''


def _impl_condition(node):
    op = node.properties.get("operator", "==")
    val = node.properties.get("value", "")
    return f'''def condition(value1, value2=None):
    """Evaluate a comparison condition (operator={op!r})."""
    _op = {op!r}
    _rhs = value2 if value2 is not None else {val!r}
    try:
        if _op == "==":     result = value1 == _rhs
        elif _op == "!=":   result = value1 != _rhs
        elif _op == ">":    result = value1 > _rhs
        elif _op == ">=":   result = value1 >= _rhs
        elif _op == "<":    result = value1 < _rhs
        elif _op == "<=":   result = value1 <= _rhs
        elif _op == "in":   result = _rhs in (value1 if hasattr(value1, "__contains__") else [])
        elif _op == "not in": result = _rhs not in (value1 if hasattr(value1, "__contains__") else [])
        else:               result = False
    except TypeError:
        result = False
    print(f"Condition: {{value1!r}} {{_op!r}} {{_rhs!r}} => {{result}}")
    return {{"true": result, "false": not result}}'''


def _impl_google_drive(node):
    folder = node.properties.get("folder_id", "")
    return f'''def google__drive__uploader(file_path):
    """Upload a file to Google Drive (requires GOOGLE_CREDS_JSON env var)."""
    import json as _json
    creds_path = os.getenv("GOOGLE_CREDS_JSON", "credentials.json")
    folder_id = {repr(folder)}
    try:
        from googleapiclient.discovery import build
        from googleapiclient.http import MediaFileUpload
        from google.oauth2.service_account import Credentials
        creds = Credentials.from_service_account_file(
            creds_path, scopes=["https://www.googleapis.com/auth/drive"])
        service = build("drive", "v3", credentials=creds)
        meta = {{"name": os.path.basename(file_path)}}
        if folder_id:
            meta["parents"] = [folder_id]
        media = MediaFileUpload(file_path)
        f = service.files().create(body=meta, media_body=media, fields="id").execute()
        print(f"Google Drive: uploaded {{repr(file_path)}} -> {{f.get('id')}}")
        return {{"file_id": f.get("id", "")}}
    except Exception as _e:
        print(f"Google Drive error: {{_e}}")
        return {{"file_id": ""}}'''


_BUILTIN_IMPLS = {
    "File Trigger":           _impl_file_trigger,
    "Excel Reader":           _impl_excel_reader,
    "File Output":            _impl_file_output,
    "Email Sender":           _impl_email_sender,
    "Condition":              _impl_condition,
    "Google Drive Uploader":  _impl_google_drive,
}


class CodeGenerator:
    """Generates executable Python code from workflow nodes using LLM."""
    
    def __init__(self):
        # Get model from environment variable or use default
        self.model = os.getenv('OLLAMA_MODEL', 'minimax-m3:cloud')
        
    def generate_workflow_code(self, workflow_data: Dict[str, Any]) -> str:
        """Generate Python code for the entire workflow."""
        nodes = [Node.from_dict(node_data) for node_data in workflow_data.get("nodes", [])]
        connections = workflow_data.get("connections", [])
        
        # Build dependency graph
        graph = self._build_dependency_graph(nodes, connections)
        
        # Generate code for each node in topological order
        node_codes = {}
        for node_id in self._topological_sort(graph):
            node = next((n for n in nodes if n.node_id == node_id), None)
            if node:
                node_codes[node_id] = self._generate_node_code(node)
        
        # Generate main workflow execution function
        main_code = self._generate_main_workflow(nodes, connections, node_codes)
        
        # Combine all code
        full_code = self._assemble_full_code(node_codes, main_code)
        
        return full_code
        
    def _build_dependency_graph(self, nodes: List[Node], connections: List[Dict[str, str]]) -> Dict[str, List[str]]:
        """Build a dependency graph from nodes and connections."""
        graph = {node.node_id: [] for node in nodes}
        
        for conn in connections:
            source_id = conn["source_id"]
            target_id = conn["target_id"]
            if target_id in graph:
                graph[target_id].append(source_id)
                
        return graph
        
    def _topological_sort(self, graph: Dict[str, List[str]]) -> List[str]:
        """Perform topological sort on the dependency graph."""
        visited = set()
        temp_visited = set()
        result = []
        
        def visit(node_id):
            if node_id in temp_visited:
                raise Exception("Cycle detected in workflow")
            if node_id not in visited:
                temp_visited.add(node_id)
                for dep_id in graph.get(node_id, []):
                    visit(dep_id)
                temp_visited.remove(node_id)
                visited.add(node_id)
                result.append(node_id)
                
        for node_id in graph:
            if node_id not in visited:
                visit(node_id)
                
        return result
        
    def _generate_node_code(self, node: Node) -> str:
        """Generate Python code for a single node — builtin first, then LLM, then stub."""
        builtin = _BUILTIN_IMPLS.get(node.title)
        if builtin:
            return builtin(node)

        prompt = self._create_node_prompt(node)
        try:
            # Call Ollama to generate code
            response = ollama.generate(
                model=self.model,
                prompt=prompt,
                options={
                    "temperature": 0.2,
                    "top_p": 0.9,
                    "stop": ["```", '"""']
                }
            )
            
            # Extract code from response
            code = response['response'].strip()

            # Strip markdown fences
            if code.startswith("```python"):
                code = code[9:].strip()
            elif code.startswith("```"):
                code = code[3:].strip()
            if code.endswith("```"):
                code = code[:-3].strip()

            # Replace JSON literals with Python equivalents
            code = _py_sanitize(code)

            return code
        except Exception as e:
            # Fallback to template-based generation
            return self._generate_fallback_code(node)
            
    def _create_node_prompt(self, node: Node) -> str:
        """Create a prompt for the LLM to generate code for a node."""
        prompt = f"""Generate Python code for a workflow node with the following specifications:

Node Type: {node.node_type.value}
Node Title: {node.title}
Inputs: {json.dumps(node.inputs, indent=2)}
Outputs: {json.dumps(node.outputs, indent=2)}
Properties: {json.dumps(node.properties, indent=2)}

Requirements:
1. Create a Python function that implements this node's functionality
2. The function name should be snake_case based on the node title
3. Function parameters should match the inputs
4. Return values should match the outputs
5. Use the properties for configuration
6. Include appropriate error handling
7. Add detailed comments explaining the code
8. Import any necessary libraries
9. Return the output values as a dictionary matching the outputs specification

Example format:
def node_function_name(input1, input2):
    # Implementation here
    return {{"output1": result1, "output2": result2}}

Generate only the Python function code without any additional text or markdown formatting:
"""
        return prompt
        
    def _generate_fallback_code(self, node: Node) -> str:
        """Generate fallback code when LLM fails."""
        func_name = self._to_snake_case(node.title.replace(" ", "_"))
        
        # Create parameter list
        params = [inp["name"] for inp in node.inputs]
        params_str = ", ".join(params) if params else ""
        
        # Build return dict — use property values where output name matches a property
        # so File Trigger with path="foo.xlsx" returns {"file_path": "foo.xlsx"} etc.
        prop_aliases = {
            "file_path": ["path", "file_path"],
            "path":      ["path", "file_path"],
            "data":      ["data"],
            "status":    ["status"],
            "file_id":   ["file_id", "folder_id"],
        }
        return_pairs = []
        for out in node.outputs:
            candidates = prop_aliases.get(out["name"], [out["name"]])
            val = None
            for c in candidates:
                if node.properties.get(c):
                    val = node.properties[c]
                    break
            return_pairs.append(f'"{out["name"]}": {repr(val)}')

        returns_str = "{" + ", ".join(return_pairs) + "}" if return_pairs else "{}"

        # Inject properties as local constants so the function body can reference them
        prop_lines = ""
        if node.properties:
            prop_lines = "\n".join(
                f'    {k} = {repr(v)}'
                for k, v in node.properties.items()
                if v not in ("", None)
            )
            if prop_lines:
                prop_lines = "\n" + prop_lines

        code = f'''def {func_name}({params_str}):
    """Stub implementation for {node.title} - replace with real logic."""{prop_lines}
    print(f"Executing {{repr('{node.title}')}}")
    return {returns_str}'''

        return code
        
    def _generate_main_workflow(self, nodes: List[Node], connections: List[Dict[str, str]], 
                               node_codes: Dict[str, str]) -> str:
        """Generate the main workflow execution function."""
        # Create execution order
        graph = self._build_dependency_graph(nodes, connections)
        execution_order = self._topological_sort(graph)
        
        # Build connection mapping
        conn_map = {}
        for conn in connections:
            source_id = conn["source_id"]
            target_id = conn["target_id"]
            if target_id not in conn_map:
                conn_map[target_id] = []
            conn_map[target_id].append(source_id)
            
        # Generate main function
        main_code = '''def execute_workflow():
    """
    Main workflow execution function
    """
    print("Starting workflow execution...")
    
    # Store results from each node
    results = {}
    
'''
        
        # Add execution for each node
        for node_id in execution_order:
            node = next((n for n in nodes if n.node_id == node_id), None)
            if not node:
                continue

            func_name = self._to_snake_case(node.title.replace(" ", "_"))

            # Build one arg per upstream connection (positional)
            connected_args = []
            for source_id in conn_map.get(node_id, []):
                source_node = next((n for n in nodes if n.node_id == source_id), None)
                if source_node:
                    if source_node.outputs:
                        out_name = source_node.outputs[0]["name"]
                        connected_args.append(f"results['{source_id}']['{out_name}']")
                    else:
                        connected_args.append(f"results['{source_id}']")

            # Pad any remaining required inputs with property value or None
            all_args = list(connected_args)
            for i in range(len(connected_args), len(node.inputs)):
                inp = node.inputs[i]
                prop_val = node.properties.get(inp["name"])
                all_args.append(repr(prop_val) if prop_val is not None else "None")

            input_str = ", ".join(all_args)

            if connected_args:
                main_code += f"    # Inputs from connected nodes: {', '.join(connected_args)}\n"

            main_code += f"    # Execute {node.title}\n"
            main_code += f"    results['{node_id}'] = {func_name}({input_str})\n"
            main_code += "\n"
            
        main_code += "    print(\"Workflow execution completed.\")\n"
        main_code += "    return results\n"
        
        return main_code
        
    def _assemble_full_code(self, node_codes: Dict[str, str], main_code: str) -> str:
        """Assemble all code into a complete Python script."""
        # Add imports
        imports = '''# -*- coding: utf-8 -*-
"""Generated FlowState Workflow"""
import os
import sys
import json
'''
        
        imports += "\n\n"
        
        # Combine all parts
        full_code = imports
        
        # Add all node functions
        for code in node_codes.values():
            full_code += code + "\n\n"
            
        # Add main function
        full_code += main_code + "\n\n"
        
        # Add execution guard
        full_code += '''if __name__ == "__main__":
    results = execute_workflow()
    print("Workflow results:", results)'''
        
        return full_code
        
    def _to_snake_case(self, name: str) -> str:
        """Convert a string to snake_case."""
        result = ""
        for i, char in enumerate(name):
            if char.isupper() and i > 0:
                result += "_"
            result += char.lower()
        return result


def generate_tests_for_workflow(workflow_data: Dict[str, Any], generated_code: str) -> str:
    """Generate pytest tests for the workflow."""
    nodes = [Node.from_dict(node_data) for node_data in workflow_data.get("nodes", [])]

    # Embed the workflow code at module level so execute_workflow is in scope
    test_code = '# -*- coding: utf-8 -*-\n'
    test_code += '"""Auto-generated tests for FlowState workflow"""\n'
    test_code += "import pytest\n"
    test_code += "from unittest.mock import patch, MagicMock\n\n"
    test_code += "# --- Workflow under test (inlined) ---\n"
    test_code += generated_code.rstrip() + "\n"
    test_code += "# --- Tests ---\n\n"
    
    # Generate tests for each node
    for node in nodes:
        func_name = node.title.replace(" ", "_").lower()
        test_code += f"def test_{func_name}():\n"
        test_code += f'    """Test for {node.title} node"""\n'
        
        # Add mock inputs based on node inputs
        if node.inputs:
            test_code += "    # Mock inputs\n"
            for inp in node.inputs:
                test_code += f"    {inp['name']} = \"mock_value\"  # Replace with actual test value\n"
            test_code += "\n"
            
        # Add function call
        params = ", ".join([inp['name'] for inp in node.inputs]) if node.inputs else ""
        test_code += f"    # TODO: Call the actual function\n"
        test_code += f"    # result = {func_name}({params})\n\n"
        
        # Add assertions based on outputs
        if node.outputs:
            test_code += "    # Assert outputs\n"
            for out in node.outputs:
                test_code += f"    # assert result['{out['name']}'] is not None\n"
                if out['type'] == 'string':
                    test_code += f"    # assert isinstance(result['{out['name']}'], str)\n"
                elif out['type'] == 'boolean':
                    test_code += f"    # assert isinstance(result['{out['name']}'], bool)\n"
                elif out['type'] == 'integer':
                    test_code += f"    # assert isinstance(result['{out['name']}'], int)\n"
                elif out['type'] == 'dict':
                    test_code += f"    # assert isinstance(result['{out['name']}'], dict)\n"
        else:
            test_code += "    # No outputs to assert\n"
            
        test_code += "\n"
        
    # Main integration test — execute_workflow is defined above (inlined)
    test_code += "def test_execute_workflow():\n"
    test_code += '    """Test the main workflow execution"""\n'
    test_code += "    result = execute_workflow()\n"
    test_code += "    assert result is not None\n"
    test_code += "    assert isinstance(result, dict)\n"

    return test_code